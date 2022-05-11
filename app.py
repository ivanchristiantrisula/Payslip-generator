import openpyxl

from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.shapes import *
import os

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


pdfmetrics.registerFont(TTFont('Arial','Arial.ttf'))

xlsxPath = input(r"Insert XLSX file path (drag n drop) : ") #USE EXCEL TEMPLATE test.xlsx

wb = openpyxl.load_workbook(xlsxPath)
sheet = wb.get_sheet_by_name('data')

#Page information
page_width = 2156
page_height = 1500
spread = 75
start = 200
start_2 = 425
start_3 = 800

#Payslip variables
company_name = 'COMPANY_NAME'
address = "COMPANY_ADDR"
month_year = input("Periode gaji : ") #PAYROLL PERIOD

def create_payslip():
    total_row = sheet.max_row
    for i in range (2, int(total_row)+1):
        try : 
            #eading values from excel file
            bank_acc_num = str(sheet.cell(row = i, column = 1).value)
            email = sheet.cell(row = i, column = 2).value
            name = sheet.cell(row = i, column = 3).value
            total_absen = sheet.cell(row = i, column = 4).value
            gaji_pokok = "Rp. "+f'{float(sheet.cell(row = i, column = 5).value or 0 ):,}'
            tunjangan = "Rp. "+f'{float(sheet.cell(row = i, column = 6).value or 0) : ,}'
            total_gaji_pokok = "Rp. "+f'{float(sheet.cell(row = i, column = 7).value or 0 ):,}'
            total_tunjangan = "Rp. "+f'{float(sheet.cell(row = i, column = 8).value or 0 ) :,}'
            claim = "Rp. "+f'{float(sheet.cell(row = i, column = 9).value or 0 ) :,}'
            bpjs = "Rp. "+f'{float(sheet.cell(row = i, column = 10).value or 0):,}'
            jam_lembur = float(sheet.cell(row = i, column = 11).value or 0 )
            biaya_lembur = "Rp. "+f'{float(sheet.cell(row = i, column = 12).value or 0 ) :,}'
            total_lembur = "Rp. "+f'{float (sheet.cell(row = i, column = 13).value or 0 ):,}'
            komisi = "Rp. "+f'{float(sheet.cell(row = i, column = 14).value or 0):,}'
            total_gaji = "Rp. "+f'{float(sheet.cell(row = i, column = 15).value or 0):,}'

            print("Generating payslip for : "+name)
            pdf_path = os.path.join("./pdf/",str(name)+'_' + month_year + '.pdf') #Generated Payslip PDF path

            #Creating a pdf file and setting a naming convention
            c = canvas.Canvas(pdf_path )
            #Page settings (size/font)
            c.setPageSize((page_width, page_height))
            c.setFont('Arial',80)

            #Company name text
            text_width = stringWidth(company_name, 'Arial',80)
            c.drawString((page_width-text_width)/2, 1400, company_name)
            
            #alamat company
            c.setFont('Arial',40)
            text_width = stringWidth(address, 'Arial',40)
            c.drawString((page_width-text_width)/2, 1325, address)
            
            #Invoice month/year information
            text = 'Slip Gaji Periode ' + month_year
            text_width = stringWidth(text, 'Arial',55)
            c.setFont('Arial',55)
            c.drawString((page_width-text_width)/2, 1150, text)

            y = 1000

            x_deduction = page_width/2+200;
            
            #Drawing payslip related information
            c.setFont('Arial',35)

            c.drawString(start, y, 'Nama')
            c.drawString(start_3, y, ": "+str(name))
            y -= spread

            
            c.drawString(start, y, 'Nomor Rekening ')
            c.drawString(start_3, y, ": "+str(bank_acc_num))
            y -= spread*2

            y_deduction = y

            c.drawString(start, y, 'Pemasukan')
            y -= spread

            c.drawString(start, y, 'Gaji Pokok')
            c.drawString(start_2,y,'( '+str(total_absen)+" x "+str(gaji_pokok)+" )")
            c.drawString(start_3, y, ": "+str(total_gaji_pokok))
            y -= spread

            c.drawString(start, y, 'Tunjangan')
            c.drawString(start_2,y,'( '+str(total_absen)+" x "+str(tunjangan)+" )")
            c.drawString(start_3, y, ": "+str(total_tunjangan))
            y -= spread

            c.drawString(start, y, 'Lembur')
            c.drawString(start_2,y,'( '+str(jam_lembur)+" x "+str(biaya_lembur)+" )")
            c.drawString(start_3, y, ": "+str(total_lembur))
            y -= spread
            
            c.drawString(start, y, 'Komisi / Bonus')
            c.drawString(start_3, y, ": "+str(komisi))
            y -= spread

            c.drawString(x_deduction, y_deduction, 'Pengeluaran')
            y_deduction -= spread

            c.drawString(x_deduction, y_deduction, 'Claim')
            c.drawString(x_deduction+500, y_deduction, ": "+str(claim))
            y_deduction -= spread

            c.drawString(x_deduction, y_deduction, 'Asuransi')
            c.drawString(x_deduction+500, y_deduction, ": "+str(bpjs))

            y -= spread

            text_width = stringWidth("Total Gaji   :   " +str(total_gaji), 'Arial', 35)
            c.drawString((page_width-text_width)/2, y,"Total Gaji   :   " +str(total_gaji))
            y -= spread * 2

            c.drawString(start, y, 'TT Direksi : ')
            c.drawString(start_2, y,'_________________')

            #draw outside border
        
            #Saving the pdf file
            c.save()

            print(name + " generated and saved!")

            if(email is not None) :
                email_payslip(pdf_path, email,month_year)
            else :
                print("")

        except Exception as e:
            print(e)
            print("Skipping...")
            print("")

def email_payslip(path,recipient_address, period) :
    print("Preparing email...")
    
    sender_address = "SENDER_EMAIL_ADDR"
    sender_password = "SENDER_EMAIL_PASSWORD"

    email_text = '''
    EMAIL_BODY
    '''
    message = MIMEMultipart();
    message["From"] = sender_address
    message["To"] = recipient_address
    message["Subject"] = "Slip Gaji COMPANY_NAME Periode " +period
    message.attach(MIMEText(email_text, 'html'))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(path, "rb").read())
    encoders.encode_base64(part)
    
    part.add_header('Content-Disposition', 'attachment; filename="slip_gaji.pdf"')

    message.attach(part)

    session = smtplib.SMTP('MAIL_SERVER', 587) #MAIL SERVER AND PORT
    session.starttls()
    session.login(sender_address,sender_password)
    text = message.as_string()
    session.sendmail(sender_address, recipient_address, text)
    session.quit()

    print("Email sent to "+recipient_address)
    print("")


create_payslip()